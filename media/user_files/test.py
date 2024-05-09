import random
from openpyxl import load_workbook
import string

list1 =['Взуття','Сорочка', 'Кофта', 'Футболка','Шорти', 'Майка','Юбка','Штани','Картуз','Окуляри','Сумка']
list2 =[]
for i in range(0,101):
    random_element = random.choice(list1)
    list2.append(random_element)

workbook = load_workbook(r"C:\Users\Bohdan\Desktop\лпну\exel\lab3\ex1.xlsx")

sheet = workbook.active

for index, value in enumerate(list2, start=1):
    sheet.cell(row=index+1, column=1, value=value)
workbook.save("ex1.xlsx")


brands = ['Adidas', 'Nike', 'Reebok', 'Puma', 'Under Armour', 'New Balance', 'Asics', 'Converse', 'Vans']

random_list = []

for i in range(101):
    brand = random.choice(brands)
    
    letter = random.choice(string.ascii_uppercase)
    
    numbers = ''.join(random.choices(string.digits, k=3))
    
    random_value = f"{brand}_{letter}_{numbers}"
    
    random_list.append(random_value)

    
for index, value in enumerate(random_list, start=1):
    sheet.cell(row=index+1, column=2, value=value)


column3 = []
for i in range(101):
    random_number = random.randint(0, 1)
    if random_number <= 0.65:
        column3.append('Продаж')
    else: column3.append('Запаси')


for index, value in enumerate(column3, start=1):
    sheet.cell(row=index+1, column=3, value=value)


column4=[]
column5=[]
column6=[]
for i in range(101):
    random_number = random.randint(1000, 4000)
    column4.append(random_number)

for i in range(101):
    random_number = random.randint(1000, 4000)
    column5.append(random_number)

for i in range(101):
    random_number = random.randint(1000, 4000)
    column6.append(random_number)
    
for index, value in enumerate(column4, start=1):
    sheet.cell(row=index+1, column=4, value=value)

for index, value in enumerate(column5, start=1):
    sheet.cell(row=index+1, column=5, value=value)

for index, value in enumerate(column6, start=1):
    sheet.cell(row=index+1, column=6, value=value)

workbook.save("ex1.xlsx")
